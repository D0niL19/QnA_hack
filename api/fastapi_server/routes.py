import numpy as np
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
import tritonclient.grpc as grpcclient
from qdrant_client import QdrantClient
from qdrant_client.http.exceptions import ResponseHandlingException
from qdrant_client.http.models import PointStruct, VectorParams
from openpyxl import load_workbook, Workbook
from pydantic import BaseModel


import pandas as pd

import fitz  # PyMuPDF
import os
import hashlib
import schedule
import time

from fastapi_server.parser import parse_main_page, parse_links

PDF_DIRECTORY = '/app/fastapi_server/example_pdf'  # Замените на ваш путь к папке с PDF
pdf_hashes = {}  # Словарь для хранения хешей PDF-файлов


# Модель данных Document
class Document(BaseModel):
    link: str
    filename: str
    title: str
    text: str

def extract_documents_from_pdf(pdf_file):
    doc = fitz.open(pdf_file)
    documents = []
    filename = pdf_file.split('/')[-1]  # Имя файла без пути

    # Извлекаем метаданные для заголовка
    metadata = doc.metadata
    title = metadata.get("title", filename)  # Используем заголовок из метаданных или имя файла


    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        text = page.get_text("text")

        # Формируем ссылку на конкретную страницу
        link = f"{page_num + 1}"

        # Создаем объект Document
        document = Document(
            link=link,
            filename=filename,
            title=title,  # Можно заменить, если найдены заголовки на странице
            text=text.strip()  # Убираем лишние пробелы
        )
        documents.append(document)

    return documents


# Шаг 2: Разбиение текста на фрагменты по количеству символов без скользящего окна
def split_text_into_chunks(text, chunk_size=512):
    words = text.split()
    chunks = []
    current_chunk = ""

    for word in words:
        # Проверка, помещается ли текущее слово в текущий чанк
        if len(current_chunk) + len(word) + 1 <= chunk_size:  # +1 для пробела
            if current_chunk:  # Если текущий чанк не пустой, добавляем пробел
                current_chunk += " "
            current_chunk += word
        else:
            if current_chunk:  # Если текущий чанк не пустой, добавляем его в список чанков
                chunks.append(current_chunk)
            current_chunk = word  # Начинаем новый чанк с текущего слова

    # Добавляем последний чанк, если он не пустой
    if current_chunk:
        chunks.append(current_chunk)

    return chunks

# Шаг 3: Подготовка данных для Qdrant (разбиение на страницы и фрагменты)
def prepare_data_for_qdrant(documents, chunk_size=512):
    data = []
    for document in documents:
        chunks = split_text_into_chunks(document.text, chunk_size)
        for chunk in chunks:
            data.append({
                "text": chunk,
                "link": document.link,
                "filename": document.filename,
                "title": document.title,
            })
    return data


def get_file_hash(file_path):
    hasher = hashlib.md5()
    with open(file_path, 'rb') as f:
        hasher.update(f.read())
    return hasher.hexdigest()

def documents_data_from_pdf():
    global pdf_hashes

    for pdf_file in os.listdir(PDF_DIRECTORY):
        if pdf_file.endswith('.pdf'):
            full_path = os.path.join(PDF_DIRECTORY, pdf_file)
            current_hash = get_file_hash(full_path)

            if pdf_file not in pdf_hashes or pdf_hashes[pdf_file] != current_hash:
                pdf_hashes[pdf_file] = current_hash  # Обновляем хеш
                documents = extract_documents_from_pdf(full_path)

                data = prepare_data_for_qdrant(documents)

                return data

# Функция для регулярного обновления в 2 часа ночи каждый день
# def schedule_updates():
#     data = documents_data_from_pdf()
#     schedule.every().day.at("02:00").do(upload_to_qdrant(data))  # Ежедневное обновление в 2 часа ночи
#     while True:
#         schedule.run_pending()
#         time.sleep(1)

def connect_to_qdrant(max_attempts=5, delay=5):
    for attempt in range(max_attempts):
        try:
            client = QdrantClient("qdrant", port=6333)
            client.get_collections()  # Проверка подключения
            return client
        except ResponseHandlingException:
            if attempt < max_attempts - 1:
                time.sleep(delay)
            else:
                raise

def find_intervals(indexies):
    numbers = []
    for i in indexies:
        numbers += [i - 2, i - 1, i, i + 1, i + 2]
    numbers = list(set(numbers))
    numbers.sort()

    intervals = []
    start = numbers[0]

    for i in range(1, len(numbers)):
        if numbers[i] != numbers[i - 1] + 1:
            intervals.append([start, numbers[i - 1]])
            start = numbers[i]

    intervals.append([start, numbers[-1]])

    return intervals


# Ensure Qdrant collection exists
def create_qdrant_collection():
    try:
        qdrant_client.get_collection("documents")
    except Exception:
         qdrant_client.create_collection(
            "documents",
            vectors_config=VectorParams(size=768, distance="Cosine")
        )


def get_embedding(text: str, model_name: str):
    input_tensors = [grpcclient.InferInput("text_input", [1], "BYTES")]
    input_tensors[0].set_data_from_numpy(np.array([text], dtype=object))

    results = triton_client.infer(model_name=model_name, inputs=input_tensors)
    embedding = results.as_numpy("text_output")[0]
    return embedding


def upload_to_qdrant(data, collection_name='documents'):
    vectors = []

    for item in data:
        text = item['text']
        embedding = get_embedding(text, "embedding")

        vectors.append(embedding)
        payload = {"text": text, "metadata": {"link": item["link"], "title": item["title"], "filename": item["filename"]}}  # Метаданные (например, номер страницы)

        collection_info = qdrant_client.get_collection(collection_name="documents")
        point = PointStruct(id=collection_info.points_count + 1, vector=embedding, payload=payload)
        qdrant_client.upsert("documents", [point])


    print("OK")


# Проверка существования файла, если нет — создаём
def init_workbook():
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        # Добавляем заголовки
        ws.append(["question", "answer", "mark", "filename"])
        wb.save(file_path)



# Функция для добавления строки
def append_data(question, answer, mark, filename):
    wb = load_workbook(file_path)
    ws = wb.active
    ws.append([question, answer, mark, filename])

    if ws.max_row % 100 == 0:
        wb.save(file_path)
    else:
        wb.save(file_path)


def update_mark(question: str, answer: str, new_mark: int):
    wb = load_workbook(file_path)
    ws = wb.active

    row_found = False
    # Ищем строку с конца в начало
    for row in reversed(list(ws.iter_rows(min_row=2, values_only=False))):  # Пропускаем заголовки
        if row[0].value == question and row[1].value == answer:
            row[2].value = new_mark  # Обновляем поле mark
            row_found = True
            break

    if row_found:
        wb.save(file_path)  # Сохраняем изменения
    else:
        raise HTTPException(status_code=404, detail="Row not found")


# # Запуск планировщика в отдельном потоке
# thread = Thread(target=schedule_updates)
# thread.start()
file_path = "/app/fastapi_server/data.xlsx"

# qdrant_client = connect_to_qdrant()
init_workbook()
qdrant_client = QdrantClient("qdrant", port=6333)
create_qdrant_collection()

router = APIRouter()
triton_client = grpcclient.InferenceServerClient(url="triton:8001")


prompt = "Ты ассистент РЖД,тебе предоставится контекст и вопрос на который тебе следует ответить, исходя из данного контекста. Напишите ответ, который отвечает на вопрос, используя только информацию, предоставленную в контексте. Дайте ответ на русском языке. Если ответа нет, то напиши только это: Я не знаю ответ на этот вопрос."


class DocumentRequest(BaseModel):
    text: str

class QuestionRequest(BaseModel):
    question: str

class AnswerResponse(BaseModel):
    answer: str

class MarkRequest(BaseModel):
    question: str
    answer: str
    mark: int

@router.post("/add_document")
def add_document(document_request: DocumentRequest):
    text = document_request.text

    pdf_file = 'fastapi_server/doc.pdf'  # Укажите путь к вашему PDF файлу

    # Извлекаем текст из PDF
    pages_text = extract_documents_from_pdf(pdf_file)
    data = prepare_data_for_qdrant(pages_text, chunk_size=512)

    upload_to_qdrant(data)

    return {"message": "Document added successfully"}

@router.post("/question", response_model=AnswerResponse)
async def ask_question(question_request: QuestionRequest):
    question = question_request.question
    question_embedding = get_embedding(question, "embedding")

    search_results = qdrant_client.search("documents", question_embedding, limit=5)

    print([f"{result.score}" for result in search_results])

    search_idx = [r.id for r in search_results]
    context_idx = find_intervals(search_idx)
    context_points = [qdrant_client.retrieve("documents", range(group[0], group[1] + 1)) for group in context_idx]

    context = ""
    for group_points in context_points:
        context += f" Название документа: {group_points[0].payload["metadata"]["filename"]}\n"
        context += f" Заголовок: {group_points[0].payload["metadata"]["title"]}\n"
        context += " ".join([point.payload["text"].strip() for point in group_points])

    print(context)

    input_tensors = [
        grpcclient.InferInput("text_input", [1], "BYTES"),
    ]

    full_request = f"Контекст:\n{context}\n\nВопрос:\n{question}\n\n"
    input_tensors[0].set_data_from_numpy(np.array([f'{prompt}____{full_request}'], dtype=object))

    results = triton_client.infer(model_name="generate", inputs=input_tensors)
    answer = results.as_numpy("text_output")[0].decode("utf-8")[:]

    append_data(question, answer, -1, f"{qdrant_client.retrieve("documents", search_idx)[0].payload["metadata"]["filename"]}")

    if answer.strip() == "":
        return AnswerResponse(answer="К сожалению, этой информации нет в моей базе данных, задайте другой вопрос и я отвечу на него")
    return AnswerResponse(answer=answer.strip())


@router.get("/download")
async def download_file():
    if os.path.exists(file_path):
        return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="data.xlsx")
    else:
        raise HTTPException(status_code=404, detail="File not found")


@router.post("/update")
async def manual_update():
    try:
        link_list = await parse_main_page()

        document_collection = await parse_links(link_list)

        data_web = document_collection.documents

        data = documents_data_from_pdf()

        upload_to_qdrant(data_web)
        # upload_to_qdrant(data)
        return {"message": "База данных обновлена"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/update_mark", response_model=dict)
async def update_data(data: MarkRequest):
    try:
        # Assuming update_mark is a function that processes the data
        update_mark(data.question, data.answer, data.mark)
        return {"status": "mark updated"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))