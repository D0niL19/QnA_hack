import os
from openpyxl import Workbook, load_workbook

file_path = "/app/fastapi_server/data.xlsx"

def init_workbook():
    """
    Инициализирует файл Excel, создавая его, если он не существует.
    В первой строке устанавливаются заголовки: "question", "answer", "mark", "filename".
    """
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["question", "answer", "mark", "filename"])
        wb.save(file_path)

def append_data(question, answer, mark, filename):
    """
    Добавляет новую строку данных в существующий файл Excel.

    Args:
        question (str): Вопрос для добавления.
        answer (str): Ответ для добавления.
        mark (int): Оценка для добавления.
        filename (str): Имя файла, связанного с вопросом и ответом.
    """
    wb = load_workbook(file_path)
    ws = wb.active
    ws.append([question, answer, mark, filename])
    wb.save(file_path)

def update_mark(question: str, answer: str, new_mark: int):
    """
    Обновляет оценку (mark) для заданного вопроса и ответа в файле Excel.

    Args:
        question (str): Вопрос, для которого необходимо обновить оценку.
        answer (str): Ответ, для которого необходимо обновить оценку.
        new_mark (int): Новая оценка для обновления.
    """
    wb = load_workbook(file_path)
    ws = wb.active
    for row in reversed(list(ws.iter_rows(min_row=2, values_only=False))):  # Пропускаем заголовки
        if row[0].value == question and row[1].value == answer:
            row[2].value = new_mark
            wb.save(file_path)
            return
